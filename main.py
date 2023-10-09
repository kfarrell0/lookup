import pandas as pd
import openpyxl
import mysql
import mysql.connector
import json
import csv
import xml.etree.ElementTree as ET


PBT = "personal bulk task.xlsx"

def lookup(file_name, value, col_find, col_give, sheet_id=0, minrow=0, maxrow=-1, colnumber=False):
    """attempts to find the passed value in the column col_find within the spreadsheet, and when it does, returns
       a list of the corresponding values in the same row within col_give. sheet_id can be either the name of a
       page in the spreadsheet, or else a 0-indexed number of that page. when colnumber is true, it uses 0-indexed
       IDs for the columns instead of the column names (like VLOOKUP). minrow is exclusive and maxrow is exclusive,
       based on a 0-indexed list of the rows starting with the first data row"""

    df = pd.read_excel(file_name, sheet_id)  # data frame of the relevant sheet

    if colnumber:
        col_find = df.columns[col_find]
        col_give = df.columns[col_give]
    cf = df[col_find]
    if maxrow < 0:
        maxrow = cf.size
    found_rows = []
    for i in range(minrow, maxrow):
        if (str(cf[i]) == str(
                value)):  # if the string version of a column value in the lookup column is the same as the
            found_rows += [i]  # string version of the passed value, take note of the row
            # print(str(cf[i]), str(value), str(cf[i] == str(value)))
    cg = df[col_give]
    ret_values = [cg[v] for v in found_rows]
    return ret_values


#returns the entire row containing value in col_find, as a dataframe
def lookup_row(file_name, value, col_find, sheet_id=0):
    df = pd.read_excel(file_name, sheet_id)
    return df[df[col_find]==value]

#returns the entire row containing value_range in col_find, assuming value_range is a list of some kind
def lookup_row_range(file_name, value_range, col_find, sheet_id=0):
    df = pd.read_excel(file_name, sheet_id)
    return df[df[col_find].isin(value_range)]


#search each sheet in the excel document given by *file_path* for cells containing *target_string*.
#the return value is a list of tuples for each sheet where it was found; the tuple contains the name of the sheet
#and a pandas dataframe of all the rows where *target_string* was found.
def search_string_in_excel(file_path, target_string):

    # Create an Excel file reader
    xls = pd.ExcelFile(file_path)

    # Initialize a DataFrame to store the results
    results = []

    # Iterate through each sheet in the Excel file
    for sheet_name in xls.sheet_names:
        # Read the sheet into a DataFrame
        df = xls.parse(sheet_name)
        # Search for the target string in the DataFrame
        filtered_df = df[df.apply(lambda row: row.astype(str).str.contains(target_string, case=False).any(), axis=1)]
        # If any rows contain the target string, add them to the results along with the sheet name
        if not filtered_df.empty:
            results.append((sheet_name, filtered_df))
    return results


def column_list(file_name, sheet_id=0):
    return pd.read_excel(file_name, sheet_id).columns


def all_column_lists(file_name):
    num_tabs = len(openpyxl.load_workbook(file_name).sheetnames)
    ret_list = []
    for i in range(num_tabs):
        ret_list += [column_list(file_name, i)]
    return ret_list

def all_col_test():
    a = all_column_lists("personal bulk task.xlsx")
    for b in a:
        print(b)


#take an excel spreadsheet with filename *src_file*, sheet name (or 0-indexed number) *sheet_id* and turn it into a csv file.
#the output filename will be the *prefix* plus the sheet_id
def export_csv(src_file, sheet_id=0, prefix=""):
    pd.read_excel(src_file, sheet_id).to_csv(prefix + str(sheet_id) + ".csv", index=False)


#this simply turns each sheet in an excel file into a separate csv file
#the file names are prefix plus the name of the sheet
def export_all_csv(src_file, prefix=""):
    wb = openpyxl.load_workbook(src_file)
    for x in wb.sheetnames:
        export_csv(src_file, x, prefix)


#get row by id
#search entire document for some string ("filter")


def lookup_test():
    print(lookup("personal bulk task.xlsx", 331, "Report 1 ID", "Second Report Name", "Commonality"))
    print([lookup("personal bulk task.xlsx", x, 1, 4, 1, colnumber = True) for x in range(331,341)])
    print(lookup("personal bulk task.xlsx", 643376, "Report ID", "Report Path", 3))
    print(lookup("personal bulk task.xlsx", 41, "Analyzer Task ID", "Folder Path", "Task Details", maxrow = 4))


def select_all(table_name):
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="ltitest000"
    )
    curs = mydb.cursor()
    query_all = "SELECT * FROM " + table_name + ";"
    curs.execute("USE test")
    curs.execute(query_all)
    ret = curs.fetchall()
    mydb.close()
    curs.close()
    return ret

def execute_query_from_file(filename):
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="ltitest000"
    )
    curs = mydb.cursor()
    query_text = open(filename, "r").read()
    curs.execute("USE test")
    curs.execute(query_text)
    ret = curs.fetchall()
    mydb.close()
    curs.close()
    return ret

#given a json file for the database details/credentials, a json file which lists relevant queries, and an identifier
#for one of the queries in the file, execute that query.
def run_query(db_file, query_file, query_id):
    db_data = json.load(open(db_file, "r"))
    mydb = mysql.connector.connect(
        host = db_data.get("host"),
        user = db_data.get("user"),
        password = db_data.get("password")
    )
    curs = mydb.cursor()
    query_json = open(query_file, "r")
    data = json.load(query_json)
    query = data.get(query_id)
    curs.execute("USE " + db_data.get("db"))
    curs.execute(query)
    results = curs.fetchall()

    output_csv = f"output_{query_id}.csv"

    with open(output_csv, 'w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        csv_writer.writerows(results)
        return f"Query results have been written to {output_csv}."

    curs.close()
    mydb.close()



def make_json_test():
    q_dict = {
        "Query1": "SELECT * FROM article_lookup WHERE Article_id=144940",
        "Query2": "SELECT * FROM article_lookup WHERE Article_id>144940",
        "Query3": "SELECT * FROM article_lookup WHERE Article_id<144940"
    }
    json_txt = json.dumps(q_dict)
    open("queries.json", "w").write(json_txt)



def parse_xml_file():
    # Load the XML file
    tree = ET.parse('sample.xml')
    root = tree.getroot()

    # Extract information based on the name of alias

    alias_name = 'eFashion'  # Replace this with the desired alias name
    for alias_element in root.findall(f"./alias[@name='{alias_name}']"):

        path = alias_element.find('path').text
        connection = alias_element.find('connection').text
        user = alias_element.find('user').text
        password = alias_element.find('pwd').text

        print(f"Alias: {alias_name}")
        print(f"Path: {path}")
        print(f"Connection: {connection}")
        print(f"User: {user}")
        print(f"Password: {password}")




if __name__ == '__main__':


    #make_json_test()
    #print(run_query("test_creds.json", "queries.json", "Query1"))
    #print(run_query("test_creds.json", "queries.json", "Query3"))
    parse_xml_file()

    #print(select_all("article_lookup")[0:4])
    #print(execute_query_from_file("select.txt"))



    #execute_query()

    #print(mysql)
    #curs.execute("CREATE DATABASE test")
    #curs.execute("USE test")
    #curs.execute("CREATE TABLE things (Name varchar(63), Number int);")
    #curs.execute("INSERT INTO things VALUES ('kevin', 1000);")
    #curs.execute("INSERT INTO things VALUES ('kanishak', 1002);")
    #curs.execute("INSERT INTO things VALUES ('missy elliot', 5);")
    #curs.execute("INSERT INTO things VALUES ('dogton blanchard', 50);")
    #mydb.commit()
    #curs.execute("SELECT * FROM things WHERE Number>100;")
    #print(curs.fetchall())







    #export_all_csv(PBT, "biswabir")
    #row_test = lookup_row(PBT, 331, "Report 1 ID", "Commonality")
    #print(lookup_row_range(PBT, range(330, 338), "Report 1 ID", "Commonality"))

    #print(row_test)
    #print(row_test["First Report Name"])
    #print(search_string_in_excel(PBT, "Calendar_year_lookup")[1])
    #export_all_csv("personal bulk task.xlsx", prefix="testhello")

    #lookup_test()
    #all_col_test()

    #for sheet in (openpyxl.load_workbook("personal bulk task.xlsx").sheetnames):
    #    print(sheet + ":")
    #    cols = column_list("personal bulk task.xlsx", sheet)
    #    if("Report ID" in cols and "Report Name" in cols):
    #        print(" ", lookup("personal bulk task.xlsx", 632214, "Report ID", "Report Name", sheet))
    #    else:
    #        print("  no relevant columns")

#dataframe.drop_duplicates(): filter out rows which are identical
#dataframe.drop_duplicates(col_name): filter out rows which have duplicate entries in col_name